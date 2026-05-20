"""
VL Construtora – Análise de Contrato
=====================================
Sistema de gestão e rastreio de solicitações de serviços por obra.

Regras de negócio
─────────────────
• 1 solicitação = 1 linha (múltiplos serviços agrupados).
• Prazo de resposta conta a partir da segunda-feira seguinte à data da solicitação.
• Fim de Cotação      = Início de Cotação + 15 dias.
• Elaboração Contrato = Fim de Cotação    + 5  dias.
• Prazo de Assinatura = Elab. Contrato   + 10 dias.
• Status automático: "Concluído" se Contrato Assinado == "Assinado".

Atualização semanal
───────────────────
Uma planilha por obra. O upload preserva dados manuais já preenchidos.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import os, base64
from datetime import datetime, date, timedelta
from io import BytesIO

# ─────────────────────────────────────────────────────────
#  LOGO em base64 (embutida para portabilidade)
# ─────────────────────────────────────────────────────────
LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABRAHgDASIAAhEBAxEB/8QAHAABAAIDAQEBAAAAAAAAAAAAAAYHAwUIBAIB/8QAShAAAQIEAgYDCggMBwAAAAAAAQIDAAQFEQYhBxITMYGRFTJBFBciI0JRUlaU0iRVcZOxsrPRCBYzNTY3YXN0dYKhQ2JydsHC8P/EABsBAAIDAQEBAAAAAAAAAAAAAAAGBAUHAgMI/8QANxEAAQMBBAQMBQUBAAAAAAAAAQACBAMFESExBkFx0RITFBYiMlFTgaGxwVJhkeHwBxWSovFC/9oADAMBAAIRAxEAPwCu9GtVbrmEpWYcCFTDI2D+QvrJ7eIseMSNyXl3EFDjDS0KFilSAQYpXQfWu4sQuUp1dmZ5PgXOQcTmOYuOUXdGcWvGMWU5oyOI8VlttRDDmOaMjiNh3Fc+6VsPtUHE6hKtbOTmk7VlI3JO5SR8h/sRERi+9MdF6VwmuaaReYkDtk+co8scs+EUJDjYsvlUUFxxGB/NiebBm8rhtLj0m4H82JCEItlcre4EoZxBiaVp6gdhfaPkdjaczz3cY6LlJGTlJdEvKyrLLSBZKEIAAEQLQdQ+4qG7WHkWenTZu/Y2k/8AJvyEWLCDb0015JptPRbh469yzfSOeZEo02nosw8de7wWpxTUZah0CbqbjbfiWyUAgeEs5JHO0VjoUw6jEFfm67VWUzDEuq4Dgulbyje5HbYZ8RGfTrWlPzkph6WUValnnkp7VnJCeVzxEWho/oQw7hSTpxSA8E7R8jtcVmeW7hD7+n1j8M8fUGGe4epXjIrGzLI4QN1StgO0NH55rdpZZSAEtNgDcAkRAtN9bbpGFO4WAhM1USWhYC4bHXP0DjFgxzVpXr/T+MZlxtetKyp7nYtuISczxNz8lo0e3pYjRCG9Z2A9/JVOilnmbPDndVnSPsPr6KJQhCM5WyJCEIELNIzL0lOszcusodZcDiFDsINxHTtCqLVWo8pUmOpMNBdvMe0cDcRy5FwaBq1tZOaoTy/CZO3YB9E5KHA2PGF3SOJxtAVhm30KV9KYXGxhWbmz0Ks51CHW1NuJCkLBSoHcQd4jmfF9IXQ8RzlNUDqtuEtk9qDmk8jHTUVbp5ou0lpSvMozaOwfI9E5pPO44iKXR6XxMnizk711blQ6MzeIlcU7J+Hjq3KoY2GHaW9Wa3KUxi+s+4Ek+ineTwFzGvi2dA9DymsQPo3+Il7jipX0DnDfaMsRI7qmvVtTtakwQ4rquvVtOStKSlmZOTZlJdAQyygNoSOwAWEfFTnGKfTpiemVarLDZcWf2AXj0RWmnWudz0yXobK7OTR2r1jubSchxP1YzyFGdLkNp9px2a1mcCK6bJbS7Tjs1qP6L5B/F2kV6tzydZqXcM07fMa1/Fp4fQmL9iG6H6B0Fg5hTqNWanfhD194BHgp4C3EmJlH0zYMEQ4bW3XE47h4BQNJbQEycQzqM6LfD7+VyimlSv8A4v4Pmn2l6s1MDYS9jmFKGZ4C55RzRFg6c6/0piro1ld5anAt5HIuHrnhkOBivoUbfmcplloybgPdaJonZvIoAc4dJ+J2ah9MfFIQhFImdIQhAhI3GDawuhYlk6kCdRtyzoHag5KHKNPCOKlNtRhY7I4Lzq021WFjsiLl1e2tLiEuIUFJUAUkbiI8WIKa1V6LN0162pMNFF/RPYeBsYjWh+tdLYSbl3V60xInYLvvKfIPLLhEzjMK9J8SuWa2n/CskkUnw5BZraf8K5bTTZtVYFJDR7rL+w1P897W5x0rh+mM0eiylMY6ku2E39I9p4m5jQN4QaTpHXiTVRsCxrJT27fqk2/05/KYl0Wls2mJgptbkBedv2VvbtrCc2m1mQF52nd7r5dWhptTjiglCAVKUdwA3mKOo7LmP9KRedSVSYc2iwfJYRuTxyH9RidaZq50XhcyLK7TE+S2LHMNjrn6BxjJoGoHR2Gl1d9FpioKui4zDSerzNzyhp0CsflNfjXjD2Gf1OC849T9ts2rN/7f0W+5/OxWMAAAALARp8a1pvD+GZ2qLI12m7NJPlOHJI5/2jcxSX4QVf7oqUth9hfi5YbZ+x3uKHgjgnP+qNmtWZyOK6oM8htP5eliwbONoTmUj1czsG/LxVWvOuPPLedWVuOKKlqO8k5kx8QhGXHFbmBdgEhCECEhCECEhCECFNND1a6Kxa3LurtLzw2C/MFeQeeXGL8jlFtam3EuIUUqSQUkbwRHS+Dauiu4bk6kCNdxuzoHYsZKHMQnaTROC9sga8D7fnySNpZC4L2yW68Dt1eXotvCERnSZXOgsJzLza9WZfGwY84UreeAueULVCi6tUbTbmTclSPQdXqtpMzJuVZYjcex1pObp0qsqlku9ztqG5Lac1r+seUdASrDUrLNSzCAhppAQhI3BIFgIqn8HugbOWm8RPo8J3xEuT6IPhnibDgYtuPpHRSzWwoQIGfoMvrmu9K5bXV2w6XUpC7x17tt68tVnEU+mzM84ha0sNqWUoBKlWG4Adpjl6rsVyqVSZqM1Tp1T0w4pxfiFbyd27dHVcIsLUsr9w4IL7gPkothW6LI4ZbT4Rdrvuw7MlyV0TVfiyd+YV90Oiar8WTvzCvujrWEVHNRnen6fdMPP6p3A/l9lyV0TVfiyd+YV90Oiar8WTvzCvujrWEHNRnen6fdHP6p3A/l9lyA806y4W3m1trG9K0kEcDCL3080FqdwwKw0ynuqSWNdYGZaORB89iQecIW7SgOg1+KJvGYKdLFtVlqRRXaLjfcRncVQsIQivVskTTR5jpWFpSZlHpNc2w6sOISlzV1FWsew78uUQuEeEiNTk0zTqC8FR5UWlKpmlVF4Kt7vxSvxE97QPdiFaQ8XuYrnJZaZdUtLS6CEtFetdROauzzAcIi0ZJdKVzDaFmyVKAJ8wvEWPZMSNUFSm24j5lQ4tjQ4tQVaTLiPmT6lWvQtLVNpFHlKZLYeeDUs0Gx8JGdt56u8m54x7e/XKer7/tI92ImqSZmsTdFv4fYYkZapJYbebQUFSbKs2o+Xr6oN9++2+P2ksyE01L1KrUSXQNhOh5llrZBaG0JIUB2KBUoX/YPNDS23pzQGtfgPkNygP0Usuo4vdTJJx6zt6lffrlPV9/2ke7Dv1ynq+/7SPdiKrplMo7NPLrDL6XZV9xmcEttkEFxOzcWgbxq3TY7id2UYUBUgK6h+m0Z5TUs3NMLEmCBrqbAICs0jVV1TuJjrnDP+PyG5c80bJ7v+zt6mHfrlPV9/wBpHuw79cp6vv8AtI92IvOydMGF1LMvIqW3TJd0tNyuq+hxy1nS52pvv+UC0eOoyUvT3sTTIpzCe5ZpnuUONXQAVnIA5EFPZ5oOcM/4/Ibkc0bJ7v8As7epp365T1ff9pHuw79cp6vv+0j3YitX2CJ2rPMUenFyQlZdTDSZUao2gb11KSOta5tfIXj2ydMpyJl9vYUyTdmJqVSEzUoXUpU4xrqaT6PhH6BeDnDP+PyG5HNGye7/ALO3rZ1bS9T6lS5qnv4eeLUw0ppXwkbiLejCI8xh6UmqFVHJem7N2bdedkQ64A6w2yTZGqTclXhjK/UhEWvakiQQapBI+Q3KdEsOHDaW0AWg9jnb1AoQhFerdIQhAhIQhAhWbjf8z4V/eN/VTGXHv6SvfyWY/wC0IQIWPR5+UoX8DNfbRpMQfnzFv7sfaNwhAhfjfXf/ANuJ+qmNlpD/AEEoPyI+zhCBC+sLfrOc/g0/YojV138vP/z5P0KhCBC9tZ/WPS//AH+I5CEIL//2Q=="

# ─────────────────────────────────────────────────────────
#  CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="VL Construtora · Análise de Contrato",
    page_icon="data:image/png;base64," + LOGO_B64[:20],
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────
#  CSS GLOBAL
# ─────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    background-color: #080d17;
    color: #d8e6f8;
    font-family: 'Inter', system-ui, sans-serif;
}
[data-testid="stAppViewContainer"] { background-color: #080d17; }
[data-testid="stHeader"]           { background: transparent; }
div.block-container {
    padding-top: 1rem; padding-bottom: 2rem;
    background-color: #080d17; max-width: 1440px;
}

/* ─── Sidebar ─── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg,#04091a 0%,#071020 50%,#09152c 100%);
    border-right: 1px solid #152640;
}
section[data-testid="stSidebar"] * { color: #b0ccee !important; }
section[data-testid="stSidebar"] .stRadio > div { gap: 4px; }
section[data-testid="stSidebar"] .stRadio label {
    font-size: 13px; font-weight: 500;
    padding: 7px 10px; border-radius: 7px;
    transition: background .15s;
}
section[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(0,100,210,.18);
}

/* ─── KPI card ─── */
.kpi-card {
    background: linear-gradient(145deg,#0b1c32 0%,#0f243f 100%);
    border: 1px solid #1a3a60;
    border-radius: 14px;
    padding: 20px 22px 16px;
    position: relative; overflow: hidden;
    transition: transform .15s, box-shadow .15s;
    min-height: 108px;
}
.kpi-card:hover { transform:translateY(-3px); box-shadow:0 10px 28px rgba(0,90,200,.2); }
.kpi-card .bar {
    position:absolute; top:0; left:0; right:0; height:3px;
    border-radius:14px 14px 0 0;
}
.kpi-card .ico { position:absolute; top:16px; right:18px; font-size:28px; opacity:.16; }
.kpi-lbl { font-size:10px; font-weight:700; letter-spacing:1.5px;
           text-transform:uppercase; color:#5a8fc0; margin-bottom:9px; }
.kpi-val { font-size:32px; font-weight:800; color:#fff; line-height:1; }
.kpi-sub { font-size:11px; color:#3d72a0; margin-top:6px; }

/* ─── Sidebar metric card ─── */
.sb-metric {
    background: linear-gradient(145deg,#0b1c32,#0f243f);
    border: 1px solid #1a3a60;
    border-radius: 11px;
    padding: 14px 16px;
    margin-bottom: 10px;
    position: relative; overflow: hidden;
}
.sb-metric .bar { position:absolute; top:0; left:0; right:0; height:3px; border-radius:11px 11px 0 0; }
.sb-metric-lbl { font-size:9px; font-weight:700; letter-spacing:1.4px;
                 text-transform:uppercase; color:#4a7aaa; margin-bottom:5px; }
.sb-metric-val { font-size:22px; font-weight:800; color:#fff; line-height:1; }
.sb-metric-sub { font-size:10px; color:#3060a0; margin-top:4px; }

/* ─── Status chip ─── */
.status-chip {
    background: linear-gradient(145deg,#0b1c32,#0f243f);
    border: 1px solid #1a3a60;
    border-radius: 12px;
    padding: 15px 16px;
    text-align: center;
    position: relative; overflow: hidden;
}
.status-chip .bar { position:absolute; top:0; left:0; right:0; height:3px; border-radius:12px 12px 0 0; }
.chip-lbl { font-size:9px; font-weight:700; letter-spacing:1.2px; text-transform:uppercase; margin-bottom:5px; }
.chip-val { font-size:26px; font-weight:800; color:#fff; }

/* ─── Section title ─── */
.sec-title {
    font-size:10px; font-weight:700; letter-spacing:2px;
    text-transform:uppercase; color:#3d72a0;
    margin-bottom:12px; padding-bottom:7px;
    border-bottom:1px solid #152640;
}

/* ─── Chart box ─── */
.chart-box {
    background: linear-gradient(145deg,#0b1c32,#0f243f);
    border: 1px solid #1a3a60;
    border-radius: 14px;
    padding: 16px 16px 8px;
    height: 100%;
}

/* ─── Upload ─── */
[data-testid="stFileUploader"] {
    background:#0b1c32; border:1.5px dashed #1a3a60; border-radius:12px; padding:8px;
}

/* ─── DataEditor ─── */
[data-testid="stDataFrame"],
[data-testid="data-editor-container"] { border:1px solid #1a3a60; border-radius:10px; overflow:hidden; }

/* ─── Buttons ─── */
.stButton > button {
    background: linear-gradient(90deg,#004db5,#0068d6);
    color:#fff; border:none; border-radius:9px; height:40px;
    font-size:13px; font-weight:600; letter-spacing:.3px; width:100%;
    transition:all .2s;
}
.stButton > button:hover { background:linear-gradient(90deg,#0068d6,#00a8e0);
    box-shadow:0 4px 14px rgba(0,104,214,.4); }
.stDownloadButton > button {
    background:linear-gradient(90deg,#04522a,#067844);
    color:#fff; border:none; border-radius:9px; height:40px;
    font-size:13px; font-weight:600; width:100%;
}

/* ─── Tabs ─── */
.stTabs [data-baseweb="tab-list"] { background:transparent; border-bottom:1px solid #1a3a60; }
.stTabs [data-baseweb="tab"] {
    background:transparent; border-radius:8px 8px 0 0;
    color:#5a8fc0; font-size:13px; font-weight:500; padding:8px 18px;
}
.stTabs [aria-selected="true"] { background:#0f243f !important;
    color:#fff !important; border-bottom:2px solid #0068d6 !important; }

hr { border-color:#152640; }
.ts-badge { background:#0b1c32; border:1px solid #1a3a60; border-radius:8px;
    padding:7px 14px; font-size:11px; color:#5a8fc0; display:inline-block; }

/* ─── Page header ─── */
.page-header {
    background: linear-gradient(135deg,#0b1c32 0%,#0e2140 60%,#0a1a32 100%);
    border: 1px solid #1a3a60;
    border-radius: 16px;
    padding: 22px 28px;
    margin-bottom: 20px;
    position: relative; overflow: hidden;
}
.page-header::before {
    content:''; position:absolute; top:0; left:0; right:0; height:3px;
    background: linear-gradient(90deg,#1adf6a,#4afa8a,#0068d6,#00a8e0);
    border-radius:16px 16px 0 0;
}
.page-title { font-size:26px; font-weight:800; color:#eaf2ff; letter-spacing:-.3px; }
.page-subtitle {
    font-size:13px; color:#4a80b0; margin-top:4px; font-weight:400; letter-spacing:.3px;
}
.page-obra-tag {
    display:inline-block; background:rgba(0,104,214,.18);
    border:1px solid rgba(0,104,214,.4); border-radius:20px;
    padding:3px 12px; font-size:11px; font-weight:600;
    color:#60a8e0; margin-top:8px; letter-spacing:.5px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  CONSTANTES / REGRAS EDITÁVEIS
# ─────────────────────────────────────────────────────────
HISTORICO_PATH            = "historico_vl.csv"
DIAS_JANELA_COTACAO       = 15   # Fim de cotação = Início cotação + N dias
DIAS_COTACAO_CONTRATO     = 5    # Elaboração contrato = Fim cotação + N dias
DIAS_CONTRATO_ASSINATURA  = 10   # Prazo assinatura = Elab contrato + N dias

COLUNAS_MANUAIS = [
    "Data Início Desejado", "Início de Cotação",
    "Contrato Assinado", "Status Cotação",
    "Observação 1", "Observação 2", "Observação 3",
]

COLUNAS_BASE2 = [
    "Solicitação", "Data solicitada", "Obra", "Serviços", "Qtd Serviços",
    "Solicitante", "Data Início Desejado", "Prazo Resposta (dias)",
    "Início de Cotação", "Fim de Cotação", "Elaboração Contrato",
    "Prazo de Assinatura", "Contrato Assinado", "Status Cotação",
    "Status", "Observação 1", "Observação 2", "Observação 3",
    "Observação Original",
]

STATUS_COTACAO_OPTS = [
    "", "Aguardando decisão da Obra", "Solicitação Incompleta",
    "Em negociação", "Em cotação", "Fase de Proposta",
    "Negociação", "Fornecedor Definido",
]

COR_STATUS = {
    "Concluído":          "#10b981",
    "Em andamento":       "#3b82f6",
    "Aguardando cotação": "#f59e0b",
    "Em atraso":          "#ef4444",
}

PLOTLY_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter", size=11, color="#9ab8d8"),
    margin=dict(l=8, r=8, t=36, b=8),
    colorway=["#0068d6","#00a8e0","#50d4f8","#003da0","#38bdf8","#7dd3fc"],
)

# ─────────────────────────────────────────────────────────
#  HELPERS — datas
# ─────────────────────────────────────────────────────────

def proxima_segunda(d: date) -> date:
    """Retorna a segunda-feira imediatamente após a data d (nunca d em si)."""
    dias_ate_segunda = (7 - d.weekday()) % 7  # 0=seg,…,6=dom → dias até próx seg
    if dias_ate_segunda == 0:
        dias_ate_segunda = 7   # se já é segunda, vai para a próxima
    return d + timedelta(days=dias_ate_segunda)


def parse_date(s) -> date | None:
    if not s or str(s).strip() in ("", "nan"):
        return None
    try:
        return pd.to_datetime(str(s), dayfirst=True, errors="coerce").date()
    except Exception:
        return None


def fmt_date(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


# ─────────────────────────────────────────────────────────
#  PARSER — Base 1 (relatório ERP, uma planilha por obra)
# ─────────────────────────────────────────────────────────

def parse_base1(uploaded_file) -> pd.DataFrame:
    """
    Lê o Excel exportado do ERP e retorna um DataFrame com UMA linha por
    solicitação — independente de quantos serviços ela contenha.
    """
    from openpyxl import load_workbook
    wb  = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws  = wb.active

    solicitacoes, sol_atual, servicos, in_items = [], {}, [], False

    for row in ws.iter_rows(values_only=True):
        rv = [str(c).strip() if c is not None else "" for c in row]

        if rv[0] == "Solicitação":
            if sol_atual:
                sol_atual["servicos"]     = servicos
                sol_atual["qtd_servicos"] = len(servicos)
                solicitacoes.append(sol_atual)
            num = next((v for v in rv if v.isdigit() and len(v) >= 4), "")
            sol_atual = {"numero": num, "data": "", "solicitante": "",
                         "obra": "", "observacao": "", "total": 0.0}
            servicos  = []
            in_items  = False

        elif rv[0] == "Data" and sol_atual and not in_items:
            sol_atual["data"]        = rv[3] if len(rv) > 3 else ""
            sol_atual["solicitante"] = rv[12] if len(rv) > 12 else ""

        elif rv[0] == "Obra" and sol_atual:
            sol_atual["obra"] = rv[3] if len(rv) > 3 else ""

        elif rv[0] == "Unidade construtiva" and sol_atual:
            in_items = True

        elif rv[0] == "Observação" and sol_atual:
            sol_atual["observacao"] = " ".join(v for v in rv if v and v != "Observação")
            in_items = False

        elif in_items and rv[0] and rv[0][0].isdigit() and "." in rv[0]:
            desc = rv[1] if len(rv) > 1 else ""
            if desc and desc not in servicos:
                servicos.append(desc)
            for v in rv:
                try:
                    val = float(v.replace(".", "").replace(",", "."))
                    if val > sol_atual.get("total", 0):
                        sol_atual["total"] = val
                except Exception:
                    pass

    if sol_atual:
        sol_atual["servicos"]     = servicos
        sol_atual["qtd_servicos"] = len(servicos)
        solicitacoes.append(sol_atual)

    rows = []
    for s in solicitacoes:
        rows.append({
            "Solicitação":          s.get("numero", ""),
            "Data solicitada":      s.get("data", ""),
            "Obra":                 s.get("obra", ""),
            "Serviços":             "; ".join(s.get("servicos", [])),
            "Qtd Serviços":         s.get("qtd_servicos", 0),
            "Solicitante":          s.get("solicitante", ""),
            "Observação Original":  s.get("observacao", ""),
            "Data Início Desejado": "", "Início de Cotação": "",
            "Contrato Assinado":    "", "Status Cotação":    "",
            "Observação 1": "",  "Observação 2": "",  "Observação 3": "",
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["Solicitação"] = df["Solicitação"].astype(str)
    return df


# ─────────────────────────────────────────────────────────
#  CÁLCULOS AUTOMÁTICOS (fórmulas da Base 2)
# ─────────────────────────────────────────────────────────

def calcular_derivados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica todas as fórmulas encadeadas da Base 2.
    Prazo Resposta parte da segunda-feira seguinte à data solicitada.
    """
    if df.empty:
        return df
    df   = df.copy()
    hoje = date.today()

    def prazo_resp(data_str):
        d = parse_date(data_str)
        if not d:
            return ""
        inicio_contagem = proxima_segunda(d)   # REGRA: conta da próx. segunda
        return max((hoje - inicio_contagem).days, 0)

    def fim_cot(inicio_str):
        d = parse_date(inicio_str)
        return fmt_date(d + timedelta(days=DIAS_JANELA_COTACAO)) if d else ""

    def elab_contrato(fim_str):
        d = parse_date(fim_str)
        return fmt_date(d + timedelta(days=DIAS_COTACAO_CONTRATO)) if d else ""

    def prazo_assn(elab_str):
        d = parse_date(elab_str)
        return fmt_date(d + timedelta(days=DIAS_CONTRATO_ASSINATURA)) if d else ""

    def calc_status(row):
        if str(row.get("Contrato Assinado", "")).strip().lower() == "assinado":
            return "Concluído"
        if not str(row.get("Início de Cotação", "")).strip():
            return "Aguardando cotação"
        prazo = row.get("Prazo Resposta (dias)", 0)
        if isinstance(prazo, (int, float)) and prazo > 30:
            return "Em atraso"
        return "Em andamento"

    df["Prazo Resposta (dias)"] = df["Data solicitada"].apply(prazo_resp)
    df["Fim de Cotação"]        = df["Início de Cotação"].apply(fim_cot)
    df["Elaboração Contrato"]   = df["Fim de Cotação"].apply(elab_contrato)
    df["Prazo de Assinatura"]   = df["Elaboração Contrato"].apply(prazo_assn)
    df["Status"]                = df.apply(calc_status, axis=1)

    for col in COLUNAS_BASE2:
        if col not in df.columns:
            df[col] = ""
    return df[COLUNAS_BASE2]


# ─────────────────────────────────────────────────────────
#  PERSISTÊNCIA
# ─────────────────────────────────────────────────────────

def carregar() -> pd.DataFrame:
    if os.path.exists(HISTORICO_PATH):
        df = pd.read_csv(HISTORICO_PATH, dtype=str).fillna("")
        for col in COLUNAS_BASE2:
            if col not in df.columns:
                df[col] = ""
        return df
    return pd.DataFrame(columns=COLUNAS_BASE2)


def salvar(df: pd.DataFrame):
    df.to_csv(HISTORICO_PATH, index=False)


def merge(historico: pd.DataFrame, novos: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Adiciona apenas registros novos; preserva campos manuais existentes."""
    if historico.empty:
        return calcular_derivados(novos), len(novos)
    existentes = set(historico["Solicitação"].astype(str))
    novas      = novos[~novos["Solicitação"].astype(str).isin(existentes)].copy()
    if novas.empty:
        return historico, 0
    resultado = pd.concat([historico, calcular_derivados(novas)], ignore_index=True)
    return calcular_derivados(resultado), len(novas)


# ─────────────────────────────────────────────────────────
#  HELPERS VISUAIS
# ─────────────────────────────────────────────────────────

def kpi(label, value, sub, icon, color="#0068d6"):
    return f"""<div class="kpi-card">
        <div class="bar" style="background:linear-gradient(90deg,{color},{color}80);"></div>
        <div class="ico">{icon}</div>
        <div class="kpi-lbl">{label}</div>
        <div class="kpi-val">{value}</div>
        <div class="kpi-sub">{sub}</div></div>"""

def chip(label, value, color):
    return f"""<div class="status-chip">
        <div class="bar" style="background:{color};"></div>
        <div class="chip-lbl" style="color:{color};">{label}</div>
        <div class="chip-val">{value}</div></div>"""

def sb_metric(label, value, sub, color="#0068d6"):
    return f"""<div class="sb-metric">
        <div class="bar" style="background:{color};"></div>
        <div class="sb-metric-lbl">{label}</div>
        <div class="sb-metric-val">{value}</div>
        <div class="sb-metric-sub">{sub}</div></div>"""


# ─────────────────────────────────────────────────────────
#  ESTADO DA SESSÃO
# ─────────────────────────────────────────────────────────
if "historico" not in st.session_state:
    raw = carregar()
    st.session_state.historico = calcular_derivados(raw) if not raw.empty else raw

historico: pd.DataFrame = st.session_state.historico


# ─────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────
with st.sidebar:
    # Logo
    st.markdown(
        f'<div style="text-align:center;padding:16px 0 6px;">'
        f'<img src="data:image/png;base64,{LOGO_B64}" style="height:64px;"/></div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div style="text-align:center;font-size:9px;letter-spacing:2.5px;'
        'color:#2d5888;padding-bottom:14px;">SISTEMA DE GESTÃO</div>',
        unsafe_allow_html=True
    )
    st.divider()

    # Navegação
    menu = st.radio("", ["📊  Análise de Contrato", "📋  Controle & Exportar"],
                    label_visibility="collapsed")
    st.divider()

    # Filtro de obra
    st.markdown(
        '<div style="font-size:9px;font-weight:700;letter-spacing:1.8px;'
        'color:#2d5888;text-transform:uppercase;margin-bottom:6px;">Filtrar Obra</div>',
        unsafe_allow_html=True
    )
    obras_disp = ["Todas as obras"]
    if not historico.empty:
        obras_disp += sorted(historico["Obra"].dropna().unique().tolist())
    obra_sel = st.selectbox("", obras_disp, label_visibility="collapsed")

    st.divider()

    # ── Métricas de tempo (sidebar) ────────────────────────
    st.markdown(
        '<div style="font-size:9px;font-weight:700;letter-spacing:1.8px;'
        'color:#2d5888;text-transform:uppercase;margin-bottom:10px;">Tempos Médios</div>',
        unsafe_allow_html=True
    )

    df_metr = historico.copy() if not historico.empty else pd.DataFrame()
    if obra_sel != "Todas as obras" and not df_metr.empty:
        df_metr = df_metr[df_metr["Obra"] == obra_sel]

    # Tempo médio solicitação → assinatura
    def media_tempo_total(df):
        vals = []
        for _, r in df.iterrows():
            d_sol  = parse_date(r.get("Data solicitada", ""))
            d_assn = parse_date(r.get("Contrato Assinado", ""))
            if d_sol and d_assn and d_assn > d_sol:
                vals.append((d_assn - d_sol).days)
        return round(sum(vals) / len(vals)) if vals else None

    def media_tempo_cotacao(df):
        """Tempo médio do início ao fim de cotação."""
        vals = []
        for _, r in df.iterrows():
            d1 = parse_date(r.get("Início de Cotação", ""))
            d2 = parse_date(r.get("Fim de Cotação", ""))
            if d1 and d2 and d2 > d1:
                vals.append((d2 - d1).days)
        return round(sum(vals) / len(vals)) if vals else None

    def media_prazo_cotacao(df):
        """Prazo médio de resposta das solicitações com cotação iniciada."""
        vals = []
        for _, r in df.iterrows():
            v = r.get("Prazo Resposta (dias)", "")
            try:
                vals.append(int(float(str(v))))
            except Exception:
                pass
        return round(sum(vals) / len(vals)) if vals else None

    tm_total  = media_tempo_total(df_metr)   if not df_metr.empty else None
    tm_cot    = media_tempo_cotacao(df_metr) if not df_metr.empty else None
    tm_prazo  = media_prazo_cotacao(df_metr) if not df_metr.empty else None

    st.markdown(sb_metric(
        "Sol. → Assinatura",
        f"{tm_total} dias" if tm_total is not None else "—",
        "tempo médio total", "#10b981"
    ), unsafe_allow_html=True)

    st.markdown(sb_metric(
        "Janela de Cotação",
        f"{tm_cot} dias" if tm_cot is not None else "—",
        "início → fim de cotação", "#3b82f6"
    ), unsafe_allow_html=True)

    st.markdown(sb_metric(
        "Prazo de Resposta",
        f"{tm_prazo} dias" if tm_prazo is not None else "—",
        "média desde a 2ª feira", "#f59e0b"
    ), unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<div style="font-size:9px;color:#1a3a60;text-align:center;padding-top:4px;">'
        'VL Construtora © 2025</div>', unsafe_allow_html=True
    )


# ── Aplica filtro global de obra ──────────────────────────
df_fil = historico.copy()
if obra_sel != "Todas as obras" and not historico.empty:
    df_fil = historico[historico["Obra"] == obra_sel].copy()


# ─────────────────────────────────────────────────────────
#  HEADER REUTILIZÁVEL
# ─────────────────────────────────────────────────────────
def render_header(titulo, subtitulo, obra_tag=None):
    tag_html = f'<div class="page-obra-tag">🏗️ {obra_tag}</div>' if obra_tag else ""
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    st.markdown(f"""
    <div class="page-header">
        <div class="page-title">{titulo}</div>
        <div class="page-subtitle">{subtitulo}</div>
        {tag_html}
        <div style="position:absolute;top:20px;right:24px;">
          <div class="ts-badge">🕐 {ts}</div>
        </div>
    </div>""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════
#  PÁGINA 1 — ANÁLISE DE CONTRATO (dashboard)
# ═════════════════════════════════════════════════════════
if menu == "📊  Análise de Contrato":

    obra_tag = obra_sel if obra_sel != "Todas as obras" else None
    render_header(
        "Análise de Contrato",
        "Monitoramento de solicitações · prazos · status de cotação e assinatura",
        obra_tag
    )

    # ── Upload ────────────────────────────────────────────
    st.markdown('<div class="sec-title">📤 Importar Base 1 (uma planilha por obra)</div>',
                unsafe_allow_html=True)

    col_up, col_btn = st.columns([5, 1])
    with col_up:
        uploaded = st.file_uploader(
            "Arraste o .xlsx exportado do ERP — uma obra por vez",
            type=["xlsx"], label_visibility="collapsed",
            help="Cada arquivo corresponde a uma obra. O sistema preserva dados manuais já preenchidos."
        )
    with col_btn:
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        if st.button("↻ Recalcular"):
            st.session_state.historico = calcular_derivados(historico)
            salvar(st.session_state.historico)
            st.rerun()

    if uploaded:
        with st.spinner("Processando…"):
            novos_df = parse_base1(uploaded)
        if novos_df.empty:
            st.warning("⚠️ Nenhuma solicitação encontrada. Verifique o arquivo.")
        else:
            df_merged, qtd = merge(historico, novos_df)
            st.session_state.historico = df_merged
            historico = df_merged
            df_fil = historico if obra_sel == "Todas as obras" else \
                     historico[historico["Obra"] == obra_sel]
            salvar(df_merged)
            if qtd == 0:
                st.info("ℹ️ Todas as solicitações já estão no histórico.")
            else:
                st.success(f"✅ {qtd} nova(s) solicitação(ões) importada(s)!")
                st.rerun()

    st.divider()

    # ── KPIs ─────────────────────────────────────────────
    st.markdown('<div class="sec-title">📈 Indicadores Gerais</div>', unsafe_allow_html=True)

    total_sol   = df_fil["Solicitação"].nunique()
    total_obras = df_fil["Obra"].nunique()
    total_pess  = df_fil["Solicitante"].nunique()
    total_serv  = df_fil["Qtd Serviços"].apply(
        lambda x: int(x) if str(x).lstrip("-").isdigit() else 0).sum()
    acima30     = int((df_fil["Prazo Resposta (dias)"].apply(
        lambda x: int(float(x)) if str(x).lstrip("-").isdigit() else 0) > 30).sum())

    k1,k2,k3,k4,k5 = st.columns(5)
    with k1: st.markdown(kpi("Solicitações",     total_sol,   f"{total_serv} serviços",       "📋","#0068d6"), unsafe_allow_html=True)
    with k2: st.markdown(kpi("Obras",            total_obras, "monitoradas",                  "🏗️","#0ea5e9"), unsafe_allow_html=True)
    with k3: st.markdown(kpi("Solicitantes",     total_pess,  "usuários com demandas",         "👤","#38bdf8"), unsafe_allow_html=True)
    with k4: st.markdown(kpi("Acima de 30 dias", acima30,     "prazo de resposta estourado",   "⏱️","#f59e0b"), unsafe_allow_html=True)
    with k5:
        conc = int((df_fil["Status"] == "Concluído").sum()) if not df_fil.empty else 0
        st.markdown(kpi("Contratos Firmados", conc, "status = Concluído", "✅","#10b981"), unsafe_allow_html=True)

    st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)

    # ── Status chips ─────────────────────────────────────
    st.markdown('<div class="sec-title">🚦 Status das Solicitações</div>', unsafe_allow_html=True)
    cnt = df_fil["Status"].value_counts() if not df_fil.empty else pd.Series(dtype=int)
    s1,s2,s3,s4 = st.columns(4)
    with s1: st.markdown(chip("Aguardando cotação", int(cnt.get("Aguardando cotação",0)), "#f59e0b"), unsafe_allow_html=True)
    with s2: st.markdown(chip("Em andamento",       int(cnt.get("Em andamento",      0)), "#3b82f6"), unsafe_allow_html=True)
    with s3: st.markdown(chip("Concluído",          int(cnt.get("Concluído",         0)), "#10b981"), unsafe_allow_html=True)
    with s4: st.markdown(chip("Em atraso",          int(cnt.get("Em atraso",         0)), "#ef4444"), unsafe_allow_html=True)

    st.divider()

    # ── Gráficos ─────────────────────────────────────────
    if not df_fil.empty:
        st.markdown('<div class="sec-title">📊 Análises Visuais</div>', unsafe_allow_html=True)

        # Linha 1
        g1, g2 = st.columns(2)

        # G1 — Solicitações por Obra (barras horizontais)
        with g1:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            d = df_fil.groupby("Obra")["Solicitação"].nunique().reset_index()
            d.columns = ["Obra","Qtd"]
            d = d.sort_values("Qtd")
            fig = go.Figure(go.Bar(
                x=d["Qtd"], y=d["Obra"], orientation="h",
                marker=dict(color=d["Qtd"],colorscale=[[0,"#003da0"],[1,"#00a8e0"]],showscale=False),
                text=d["Qtd"], textposition="outside",
                textfont=dict(color="#d0e8f8",size=12),
            ))
            fig.update_layout(title=dict(text="Solicitações por Obra",font=dict(size=13,color="#5a8fc0")),
                              xaxis=dict(visible=False,showgrid=False),
                              yaxis=dict(showgrid=False), **PLOTLY_BASE)
            st.plotly_chart(fig, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # G2 — Distribuição por Status (donut)
        with g2:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            if not cnt.empty:
                cores = [COR_STATUS.get(s,"#0068d6") for s in cnt.index]
                fig2 = go.Figure(go.Pie(
                    labels=cnt.index.tolist(), values=cnt.values.tolist(),
                    hole=0.62,
                    marker=dict(colors=cores),
                    textfont=dict(color="#d0e8f8",size=11),
                    hovertemplate="%{label}<br>%{value} solicitações<extra></extra>",
                ))
                fig2.update_layout(title=dict(text="Distribuição por Status",font=dict(size=13,color="#5a8fc0")),
                                   legend=dict(font=dict(color="#9ab8d8",size=10)), **PLOTLY_BASE)
                st.plotly_chart(fig2, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Linha 2
        g3, g4 = st.columns(2)

        # G3 — Evolução mensal de solicitações (área)
        with g3:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            df_t = df_fil.copy()
            df_t["dt"] = pd.to_datetime(df_t["Data solicitada"], dayfirst=True, errors="coerce")
            df_t = df_t.dropna(subset=["dt"])
            if not df_t.empty:
                ev = df_t.groupby(df_t["dt"].dt.to_period("M"))["Solicitação"].nunique().reset_index()
                ev["Mês"] = ev["dt"].astype(str)
                fig3 = go.Figure(go.Scatter(
                    x=ev["Mês"], y=ev["Solicitação"],
                    mode="lines+markers",
                    line=dict(color="#0068d6",width=2.5),
                    marker=dict(color="#00a8e0",size=7),
                    fill="tozeroy", fillcolor="rgba(0,104,214,.10)",
                    hovertemplate="Mês: %{x}<br>%{y} solicitações<extra></extra>",
                ))
                fig3.update_layout(title=dict(text="Evolução Mensal de Solicitações",font=dict(size=13,color="#5a8fc0")),
                                   xaxis=dict(showgrid=False), yaxis=dict(showgrid=False), **PLOTLY_BASE)
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("Sem dados temporais suficientes.")
            st.markdown('</div>', unsafe_allow_html=True)

        # G4 — Solicitações por Solicitante (barras verticais)
        with g4:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            ds = df_fil.groupby("Solicitante")["Solicitação"].nunique().reset_index()
            ds.columns = ["Solicitante","Qtd"]
            ds = ds.sort_values("Qtd", ascending=False)
            fig4 = go.Figure(go.Bar(
                x=ds["Solicitante"], y=ds["Qtd"],
                marker=dict(color=ds["Qtd"],colorscale=[[0,"#003da0"],[1,"#00a8e0"]],showscale=False),
                text=ds["Qtd"], textposition="outside",
                textfont=dict(color="#d0e8f8",size=12),
            ))
            fig4.update_layout(title=dict(text="Solicitações por Solicitante",font=dict(size=13,color="#5a8fc0")),
                               xaxis=dict(showgrid=False), yaxis=dict(visible=False,showgrid=False), **PLOTLY_BASE)
            st.plotly_chart(fig4, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Linha 3
        g5, g6 = st.columns(2)

        # G5 — Prazo de resposta: distribuição por faixa (histograma)
        with g5:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            prazos = pd.to_numeric(df_fil["Prazo Resposta (dias)"], errors="coerce").dropna()
            if not prazos.empty:
                fig5 = go.Figure(go.Histogram(
                    x=prazos, nbinsx=10,
                    marker=dict(color="#0068d6", line=dict(color="#00a8e0",width=1)),
                    hovertemplate="Faixa: %{x} dias<br>Qtd: %{y}<extra></extra>",
                ))
                fig5.update_layout(title=dict(text="Distribuição do Prazo de Resposta (dias)",font=dict(size=13,color="#5a8fc0")),
                                   xaxis=dict(showgrid=False,title="dias"),
                                   yaxis=dict(showgrid=False), **PLOTLY_BASE)
                st.plotly_chart(fig5, use_container_width=True)
            else:
                st.info("Sem dados de prazo para exibir.")
            st.markdown('</div>', unsafe_allow_html=True)

        # G6 — Status Cotação (donut)
        with g6:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            sc = df_fil["Status Cotação"].replace("","(sem status)").value_counts()
            if not sc.empty:
                fig6 = go.Figure(go.Pie(
                    labels=sc.index.tolist(), values=sc.values.tolist(),
                    hole=0.55,
                    marker=dict(colors=["#0068d6","#00a8e0","#50d4f8","#003da0","#38bdf8","#7dd3fc","#0a2a5a"]),
                    textfont=dict(color="#d0e8f8",size=11),
                    hovertemplate="%{label}<br>%{value}<extra></extra>",
                ))
                fig6.update_layout(title=dict(text="Status de Cotação",font=dict(size=13,color="#5a8fc0")),
                                   legend=dict(font=dict(color="#9ab8d8",size=10)), **PLOTLY_BASE)
                st.plotly_chart(fig6, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.info("📂 Nenhum dado ainda. Faça o upload da Base 1 para começar.")


# ═════════════════════════════════════════════════════════
#  PÁGINA 2 — CONTROLE & EXPORTAR (unificado)
# ═════════════════════════════════════════════════════════
elif menu == "📋  Controle & Exportar":

    obra_tag = obra_sel if obra_sel != "Todas as obras" else None
    render_header(
        "Controle & Exportar",
        "Edição de campos manuais · histórico consolidado · exportação de dados",
        obra_tag
    )

    if df_fil.empty:
        st.info("📂 Nenhum dado. Importe a Base 1 na página Análise de Contrato.")
    else:
        # ── Filtros ───────────────────────────────────────
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            st_opts = ["Todos"] + sorted(df_fil["Status"].dropna().unique().tolist())
            f_st = st.selectbox("Status", st_opts)
        with cf2:
            sl_opts = ["Todos"] + sorted(df_fil["Solicitante"].dropna().unique().tolist())
            f_sl = st.selectbox("Solicitante", sl_opts)
        with cf3:
            busca = st.text_input("🔍 Buscar", placeholder="nº solicitação ou palavra-chave…")

        df_view = df_fil.copy()
        if f_st != "Todos": df_view = df_view[df_view["Status"]     == f_st]
        if f_sl != "Todos": df_view = df_view[df_view["Solicitante"]== f_sl]
        if busca:
            mask = (df_view["Solicitação"].str.contains(busca,case=False,na=False) |
                    df_view["Serviços"].str.contains(busca,case=False,na=False))
            df_view = df_view[mask]

        st.markdown(
            f"<div style='font-size:11px;color:#3d72a0;margin:6px 0 10px;'>"
            f"{len(df_view)} registro(s) exibido(s)</div>", unsafe_allow_html=True
        )

        # ── Tabs ─────────────────────────────────────────
        tab_view, tab_edit, tab_exp = st.tabs(["👁️ Visualizar", "✏️ Editar", "📥 Exportar"])

        COLS_VIEW = [
            "Solicitação","Data solicitada","Obra","Solicitante",
            "Serviços","Qtd Serviços","Prazo Resposta (dias)",
            "Início de Cotação","Fim de Cotação",
            "Elaboração Contrato","Prazo de Assinatura",
            "Contrato Assinado","Status Cotação","Status",
        ]

        with tab_view:
            st.dataframe(
                df_view[COLS_VIEW], use_container_width=True, hide_index=True,
                column_config={
                    "Prazo Resposta (dias)": st.column_config.NumberColumn("Prazo (dias)", format="%d dias"),
                    "Qtd Serviços":          st.column_config.NumberColumn("Qtd Serv."),
                }
            )

        with tab_edit:
            st.markdown(
                "<div style='font-size:11px;color:#5a8fc0;margin-bottom:10px;'>"
                "Edite os campos manuais e clique em <strong>Salvar</strong>.</div>",
                unsafe_allow_html=True
            )
            COLS_EDIT = [
                "Solicitação","Obra","Serviços",
                "Data Início Desejado","Início de Cotação",
                "Contrato Assinado","Status Cotação",
                "Observação 1","Observação 2","Observação 3",
            ]
            df_ed = st.data_editor(
                df_view[COLS_EDIT].copy(),
                column_config={
                    "Solicitação":    st.column_config.TextColumn("Nº", disabled=True),
                    "Obra":           st.column_config.TextColumn("Obra", disabled=True),
                    "Serviços":       st.column_config.TextColumn("Serviços", disabled=True),
                    "Status Cotação": st.column_config.SelectboxColumn("Status Cotação",options=STATUS_COTACAO_OPTS),
                    "Contrato Assinado": st.column_config.SelectboxColumn("Contrato Assinado",
                                         options=["","Assinado","Pendente","Cancelado"]),
                },
                use_container_width=True, hide_index=True, num_rows="fixed",
            )
            if st.button("💾 Salvar alterações"):
                idx = df_view.index
                for col in COLUNAS_MANUAIS:
                    if col in df_ed.columns:
                        historico.loc[idx, col] = df_ed[col].values
                st.session_state.historico = calcular_derivados(historico)
                salvar(st.session_state.historico)
                st.success("✅ Salvo! Prazos recalculados.")
                st.rerun()

        # ── Exportar ──────────────────────────────────────
        with tab_exp:
            st.markdown(kpi(
                "Registros disponíveis", len(historico),
                f"{historico['Obra'].nunique()} obra(s) · "
                f"{historico['Solicitação'].nunique()} solicitação(ões)",
                "📂","#0068d6"
            ), unsafe_allow_html=True)
            st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)

            ex1, ex2 = st.columns(2)
            with ex1:
                st.markdown('<div class="sec-title">📄 CSV</div>', unsafe_allow_html=True)
                csv = historico.to_csv(index=False).encode("utf-8-sig")
                st.download_button("⬇️ Baixar CSV", data=csv,
                    file_name=f"vl_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv")
            with ex2:
                st.markdown('<div class="sec-title">📊 Excel (.xlsx)</div>', unsafe_allow_html=True)
                try:
                    buf = BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        historico.to_excel(writer, sheet_name="Controle", index=False)
                        pd.DataFrame({"Status Cotação": STATUS_COTACAO_OPTS[1:]}
                                     ).to_excel(writer, sheet_name="Referência", index=False)
                    st.download_button("⬇️ Baixar Excel", data=buf.getvalue(),
                        file_name=f"vl_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")

            st.divider()
            st.markdown('<div class="sec-title">👁️ Pré-visualização</div>', unsafe_allow_html=True)
            st.dataframe(historico.head(15), use_container_width=True, hide_index=True)
