"""
VL Construtora – Sistema de Gestão de Solicitações
===================================================
Transforma automaticamente o relatório da Base 1 (Excel exportado do ERP)
para o formato da Base 2 (planilha de controle com fórmulas de prazo).

Regras de negócio principais
──────────────────────────────
• 1 solicitação = 1 registro, independente de quantos serviços/itens ela contenha.
• Prazo de resposta = TODAY() - data da solicitação (calculado dinâmicamente).
• Fim de cotação = Início de cotação + 15 dias.
• Elaboração contrato = Fim de cotação + 5 dias.
• Prazo de assinatura = Elaboração contrato + 10 dias.
• Status automático: "Concluído" se Contrato Assinado == "Assinado", senão "Em atraso".

Atualização semanal
───────────────────
Basta fazer upload de um novo .xlsx da Base 1. O sistema detecta solicitações
novas (que ainda não estão no histórico) e as adiciona sem sobrescrever os
dados já preenchidos manualmente (Início de cotação, Contrato assinado, etc.).
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import os
from datetime import datetime, date, timedelta

# ─────────────────────────────────────────────────────────
#  CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="VL Construtora · Gestão",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────────────
#  ESTILO GLOBAL — Paleta azul corporativa Power BI Dark
# ─────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── Reset base ── */
html, body, [class*="css"] {
    background-color: #0b0f1a;
    color: #dce6f5;
    font-family: 'Inter', system-ui, sans-serif;
}

/* ── App container ── */
[data-testid="stAppViewContainer"] { background-color: #0b0f1a; }
[data-testid="stHeader"]           { background: transparent; }
div.block-container {
    padding-top: 1.2rem;
    padding-bottom: 2rem;
    background-color: #0b0f1a;
    max-width: 1400px;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #060c1a 0%, #0c1628 100%);
    border-right: 1px solid #162847;
    width: 240px !important;
}
section[data-testid="stSidebar"] * { color: #b8d0ee !important; }
section[data-testid="stSidebar"] .stRadio > label {
    font-size: 13px;
    font-weight: 500;
}

/* ── KPI Cards ── */
.kpi-card {
    background: linear-gradient(140deg, #0d1e35 0%, #102542 100%);
    border: 1px solid #1c3d63;
    border-radius: 14px;
    padding: 22px 24px 18px 24px;
    position: relative;
    overflow: hidden;
    transition: transform .15s, box-shadow .15s;
    min-height: 110px;
}
.kpi-card:hover {
    transform: translateY(-3px);
    box-shadow: 0 10px 28px rgba(0,100,210,.22);
}
.kpi-card .accent-bar {
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    border-radius: 14px 14px 0 0;
}
.kpi-card .icon {
    position: absolute;
    top: 18px; right: 20px;
    font-size: 30px;
    opacity: .18;
}
.kpi-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.4px;
    text-transform: uppercase;
    color: #6ea8d8;
    margin-bottom: 10px;
}
.kpi-value {
    font-size: 34px;
    font-weight: 800;
    color: #ffffff;
    line-height: 1;
}
.kpi-sub {
    font-size: 11px;
    color: #4d84b8;
    margin-top: 7px;
}

/* ── Status mini-cards ── */
.status-row {
    display: flex;
    gap: 14px;
    margin-bottom: 4px;
}
.status-chip {
    flex: 1;
    background: linear-gradient(140deg, #0d1e35 0%, #102542 100%);
    border: 1px solid #1c3d63;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    position: relative;
    overflow: hidden;
}
.status-chip .chip-bar {
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    border-radius: 12px 12px 0 0;
}
.chip-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    margin-bottom: 6px;
}
.chip-value {
    font-size: 28px;
    font-weight: 800;
    color: #fff;
}

/* ── Section titles ── */
.sec-title {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1.8px;
    text-transform: uppercase;
    color: #4d84b8;
    margin-bottom: 14px;
    padding-bottom: 8px;
    border-bottom: 1px solid #162847;
}

/* ── Chart wrapper ── */
.chart-box {
    background: linear-gradient(140deg, #0d1e35 0%, #102542 100%);
    border: 1px solid #1c3d63;
    border-radius: 14px;
    padding: 18px 18px 10px;
}

/* ── Upload ── */
[data-testid="stFileUploader"] {
    background: #0d1e35;
    border: 1.5px dashed #1c3d63;
    border-radius: 12px;
    padding: 8px;
}

/* ── DataEditor / DataFrame ── */
[data-testid="stDataFrame"],
[data-testid="data-editor-container"] {
    border: 1px solid #1c3d63;
    border-radius: 10px;
    overflow: hidden;
}

/* ── Botões ── */
.stButton > button {
    background: linear-gradient(90deg, #0052a8, #006fd6);
    color: #fff;
    border: none;
    border-radius: 9px;
    height: 42px;
    font-size: 13px;
    font-weight: 600;
    letter-spacing: .3px;
    width: 100%;
    transition: all .2s;
}
.stButton > button:hover {
    background: linear-gradient(90deg, #006fd6, #00b0e8);
    box-shadow: 0 4px 14px rgba(0,111,214,.4);
}
.stDownloadButton > button {
    background: linear-gradient(90deg, #045d30, #05875f);
    color: #fff;
    border: none;
    border-radius: 9px;
    height: 42px;
    font-size: 13px;
    font-weight: 600;
    width: 100%;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: transparent;
    border-bottom: 1px solid #1c3d63;
    gap: 0;
}
.stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 8px 8px 0 0;
    color: #6ea8d8;
    font-size: 13px;
    font-weight: 500;
    padding: 8px 20px;
}
.stTabs [aria-selected="true"] {
    background: #102542 !important;
    color: #fff !important;
    border-bottom: 2px solid #006fd6 !important;
}

/* ── Alerts ── */
div[data-testid="stAlert"] {
    border-radius: 10px;
    border-left: 4px solid;
}

hr { border-color: #162847; }

/* ── Timestamp badge ── */
.ts-badge {
    background: #0d1e35;
    border: 1px solid #1c3d63;
    border-radius: 8px;
    padding: 8px 16px;
    font-size: 12px;
    color: #6ea8d8;
    display: inline-block;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  CONSTANTES E CONFIGURAÇÕES EDITÁVEIS
# ─────────────────────────────────────────────────────────

# Arquivo de persistência do histórico processado
HISTORICO_PATH = "historico_vl.csv"

# ──── REGRAS DE PRAZO (editáveis aqui) ────────────────────
# Dias adicionados ao Fim de cotação para gerar Elaboração do contrato
DIAS_COTACAO_PARA_CONTRATO = 5

# Dias adicionados à Elaboração do contrato para gerar Prazo de assinatura
DIAS_CONTRATO_PARA_ASSINATURA = 10

# Duração padrão da janela de cotação (dias)
DIAS_JANELA_COTACAO = 15

# Colunas que o usuário preenche manualmente (não são sobrescritas no merge)
COLUNAS_MANUAIS = [
    "Data Início Desejado",
    "Início de Cotação",
    "Contrato Assinado",
    "Status Cotação",
    "Observação 1",
    "Observação 2",
    "Observação 3",
]

# Colunas da Base 2 (modelo de saída)
COLUNAS_BASE2 = [
    "Solicitação",
    "Data solicitada",
    "Obra",
    "Serviços",                # descrição consolidada dos serviços
    "Qtd Serviços",            # quantidade de itens na solicitação
    "Solicitante",
    "Data Início Desejado",    # manual
    "Prazo Resposta (dias)",   # calculado: hoje - data solicitação
    "Início de Cotação",       # manual
    "Fim de Cotação",          # calculado: Início cotação + 15 dias
    "Elaboração Contrato",     # calculado: Fim cotação + 5 dias
    "Prazo de Assinatura",     # calculado: Elab contrato + 10 dias
    "Contrato Assinado",       # manual
    "Status Cotação",          # manual (dropdown)
    "Status",                  # automático: baseado em Contrato Assinado
    "Observação 1",            # manual
    "Observação 2",            # manual
    "Observação 3",            # manual
    "Observação Original",     # da Base 1, somente leitura
]

# ─────────────────────────────────────────────────────────
#  PARSERS — Base 1
# ─────────────────────────────────────────────────────────

def parse_base1(uploaded_file) -> pd.DataFrame:
    """
    Lê o arquivo .xlsx da Base 1 (relatório ERP) e retorna um DataFrame
    onde cada linha representa UMA solicitação (independente de quantos
    serviços ela contenha).

    Campos extraídos por solicitação:
        numero, data, solicitante, obra, servicos (lista), observacao, total
    """
    from openpyxl import load_workbook

    wb  = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws  = wb.active

    solicitacoes = []
    sol_atual    = {}
    servicos     = []
    in_items     = False

    for row in ws.iter_rows(values_only=True):
        # Normaliza: converte None → "", demais → str stripped
        rv = [str(c).strip() if c is not None else "" for c in row]

        # ── INÍCIO DE NOVA SOLICITAÇÃO ─────────────────────────
        if rv[0] == "Solicitação":
            # Salva a solicitação anterior (se existir)
            if sol_atual:
                sol_atual["servicos"]   = servicos
                sol_atual["qtd_servicos"] = len(servicos)
                solicitacoes.append(sol_atual)

            # Novo bloco
            num = next((v for v in rv if v.isdigit() and len(v) >= 4), "")
            sol_atual = {
                "numero": num, "data": "", "solicitante": "",
                "obra": "", "observacao": "", "total": 0.0,
            }
            servicos = []
            in_items = False

        # ── DATA E SOLICITANTE ─────────────────────────────────
        elif rv[0] == "Data" and sol_atual and not in_items:
            sol_atual["data"]        = rv[3] if len(rv) > 3 else ""
            sol_atual["solicitante"] = rv[12] if len(rv) > 12 else ""

        # ── OBRA ───────────────────────────────────────────────
        elif rv[0] == "Obra" and sol_atual:
            sol_atual["obra"] = rv[3] if len(rv) > 3 else ""

        # ── UNIDADE CONSTRUTIVA (marca início dos itens) ───────
        elif rv[0] == "Unidade construtiva" and sol_atual:
            in_items = True

        # ── OBSERVAÇÃO (marca fim dos itens) ──────────────────
        elif rv[0] == "Observação" and sol_atual:
            obs_parts = [v for v in rv if v and v != "Observação"]
            sol_atual["observacao"] = " ".join(obs_parts)
            in_items = False

        # ── ITEM DE SERVIÇO ────────────────────────────────────
        elif in_items and rv[0] and rv[0][0].isdigit() and "." in rv[0]:
            descricao = rv[1] if len(rv) > 1 else ""
            if descricao and descricao not in ("", "nan"):
                # Evita duplicatas de serviço dentro da mesma solicitação
                if descricao not in servicos:
                    servicos.append(descricao)

            # Soma valor total (pega o maior número numérico da linha)
            for v in rv:
                try:
                    val = float(v.replace(".", "").replace(",", "."))
                    if val > sol_atual.get("total", 0):
                        sol_atual["total"] = val
                except Exception:
                    pass

    # Última solicitação
    if sol_atual:
        sol_atual["servicos"]     = servicos
        sol_atual["qtd_servicos"] = len(servicos)
        solicitacoes.append(sol_atual)

    # ── Monta DataFrame ────────────────────────────────────────
    rows = []
    for s in solicitacoes:
        descricao_consolidada = "; ".join(s.get("servicos", []))
        rows.append({
            "Solicitação":          s.get("numero", ""),
            "Data solicitada":      s.get("data", ""),
            "Obra":                 s.get("obra", ""),
            "Serviços":             descricao_consolidada,
            "Qtd Serviços":         s.get("qtd_servicos", 0),
            "Solicitante":          s.get("solicitante", ""),
            "Observação Original":  s.get("observacao", ""),
            # Campos calculados/manuais – iniciam vazios
            "Data Início Desejado": "",
            "Início de Cotação":    "",
            "Contrato Assinado":    "",
            "Status Cotação":       "",
            "Observação 1":         "",
            "Observação 2":         "",
            "Observação 3":         "",
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Garante que Solicitação seja string para merge consistente
    df["Solicitação"] = df["Solicitação"].astype(str)
    return df


# ─────────────────────────────────────────────────────────
#  CÁLCULOS AUTOMÁTICOS (Base 2 lógica)
# ─────────────────────────────────────────────────────────

def calcular_colunas_derivadas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica as fórmulas da Base 2 ao DataFrame:
    • Prazo Resposta = hoje - data solicitação
    • Fim de Cotação = Início cotação + DIAS_JANELA_COTACAO
    • Elaboração Contrato = Fim cotação + DIAS_COTACAO_PARA_CONTRATO
    • Prazo de Assinatura = Elab contrato + DIAS_CONTRATO_PARA_ASSINATURA
    • Status = "Concluído" se Contrato Assinado == "Assinado", senão "Em atraso"
    """
    if df.empty:
        return df

    df = df.copy()
    hoje = date.today()

    # ── Prazo de Resposta ──────────────────────────────────────
    def prazo_resposta(data_str):
        try:
            d = pd.to_datetime(data_str, dayfirst=True, errors="coerce")
            if pd.isna(d):
                return ""
            return (hoje - d.date()).days
        except Exception:
            return ""

    df["Prazo Resposta (dias)"] = df["Data solicitada"].apply(prazo_resposta)

    # ── Fim de Cotação ─────────────────────────────────────────
    def fim_cotacao(inicio_str):
        try:
            d = pd.to_datetime(inicio_str, dayfirst=True, errors="coerce")
            if pd.isna(d):
                return ""
            return (d.date() + timedelta(days=DIAS_JANELA_COTACAO)).strftime("%d/%m/%Y")
        except Exception:
            return ""

    df["Fim de Cotação"] = df["Início de Cotação"].apply(fim_cotacao)

    # ── Elaboração Contrato ────────────────────────────────────
    def elab_contrato(fim_str):
        try:
            d = pd.to_datetime(fim_str, dayfirst=True, errors="coerce")
            if pd.isna(d):
                return ""
            return (d.date() + timedelta(days=DIAS_COTACAO_PARA_CONTRATO)).strftime("%d/%m/%Y")
        except Exception:
            return ""

    df["Elaboração Contrato"] = df["Fim de Cotação"].apply(elab_contrato)

    # ── Prazo de Assinatura ────────────────────────────────────
    def prazo_assinatura(elab_str):
        try:
            d = pd.to_datetime(elab_str, dayfirst=True, errors="coerce")
            if pd.isna(d):
                return ""
            return (d.date() + timedelta(days=DIAS_CONTRATO_PARA_ASSINATURA)).strftime("%d/%m/%Y")
        except Exception:
            return ""

    df["Prazo de Assinatura"] = df["Elaboração Contrato"].apply(prazo_assinatura)

    # ── Status automático ──────────────────────────────────────
    def calcular_status(row):
        assinado = str(row.get("Contrato Assinado", "")).strip().lower()
        if assinado == "assinado":
            return "Concluído"
        inicio_cot = str(row.get("Início de Cotação", "")).strip()
        if not inicio_cot:
            return "Aguardando cotação"
        prazo_r = row.get("Prazo Resposta (dias)", "")
        if isinstance(prazo_r, (int, float)) and prazo_r > 30:
            return "Em atraso"
        return "Em andamento"

    df["Status"] = df.apply(calcular_status, axis=1)

    # Garante ordem das colunas
    for col in COLUNAS_BASE2:
        if col not in df.columns:
            df[col] = ""

    return df[COLUNAS_BASE2]


# ─────────────────────────────────────────────────────────
#  PERSISTÊNCIA
# ─────────────────────────────────────────────────────────

def carregar_historico() -> pd.DataFrame:
    """Carrega o CSV de histórico ou cria um DataFrame vazio."""
    if os.path.exists(HISTORICO_PATH):
        df = pd.read_csv(HISTORICO_PATH, dtype=str).fillna("")
        # Garante todas as colunas (retrocompatibilidade)
        for col in COLUNAS_BASE2:
            if col not in df.columns:
                df[col] = ""
        return df
    return pd.DataFrame(columns=COLUNAS_BASE2)


def salvar_historico(df: pd.DataFrame):
    """Salva o DataFrame no CSV de histórico."""
    df.to_csv(HISTORICO_PATH, index=False)


def merge_com_historico(historico: pd.DataFrame, novos: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Mescla novos registros com o histórico existente.

    Regra de merge:
    • Solicitações novas → adicionadas integralmente.
    • Solicitações já existentes → preserva os campos manuais do histórico;
      atualiza apenas campos derivados da Base 1 (Obra, Serviços, etc.).

    Retorna (df_atualizado, qtd_novas_adicionadas).
    """
    if historico.empty:
        df_calc = calcular_colunas_derivadas(novos)
        return df_calc, len(novos)

    existentes = set(historico["Solicitação"].astype(str).unique())
    novas      = novos[~novos["Solicitação"].astype(str).isin(existentes)].copy()

    if novas.empty:
        return historico, 0

    novas_calc = calcular_colunas_derivadas(novas)
    resultado  = pd.concat([historico, novas_calc], ignore_index=True)
    resultado  = calcular_colunas_derivadas(resultado)  # recalcula derivados com dados atuais

    return resultado, len(novas)


# ─────────────────────────────────────────────────────────
#  HELPERS VISUAIS
# ─────────────────────────────────────────────────────────

PLOTLY_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter", size=11, color="#b8d0ee"),
    margin=dict(l=8, r=8, t=36, b=8),
    colorway=["#006fd6", "#00b0e8", "#50d4f8", "#0042a0", "#38bdf8", "#7dd3fc"],
)

COR_STATUS = {
    "Concluído":          "#10b981",
    "Em andamento":       "#3b82f6",
    "Aguardando cotação": "#f59e0b",
    "Em atraso":          "#ef4444",
}

def kpi_card(label: str, value, sub: str, icon: str, color: str = "#006fd6") -> str:
    return f"""
    <div class="kpi-card">
        <div class="accent-bar" style="background:linear-gradient(90deg,{color},{color}aa);"></div>
        <div class="icon">{icon}</div>
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
    </div>"""

def status_chip(label: str, value: int, color: str) -> str:
    return f"""
    <div class="status-chip">
        <div class="chip-bar" style="background:{color};"></div>
        <div class="chip-label" style="color:{color};">{label}</div>
        <div class="chip-value">{value}</div>
    </div>"""


# ─────────────────────────────────────────────────────────
#  ESTADO DA SESSÃO
# ─────────────────────────────────────────────────────────

if "historico" not in st.session_state:
    df_hist = carregar_historico()
    st.session_state.historico = calcular_colunas_derivadas(df_hist) if not df_hist.empty else df_hist

historico: pd.DataFrame = st.session_state.historico


# ─────────────────────────────────────────────────────────
#  SIDEBAR — NAVEGAÇÃO
# ─────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style="padding:14px 0 24px; text-align:center;">
        <div style="font-size:42px; line-height:1;">🏗️</div>
        <div style="font-size:17px; font-weight:800; color:#e8f0ff; margin-top:8px;
                    letter-spacing:.5px;">VL CONSTRUTORA</div>
        <div style="font-size:10px; color:#4d84b8; letter-spacing:2.5px; margin-top:3px;">
            GESTÃO DE SOLICITAÇÕES</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    menu = st.radio(
        "Navegação",
        ["📊  Dashboard", "📋  Controle Detalhado", "📥  Exportar"],
        label_visibility="collapsed",
    )

    # ── Filtro de obra no sidebar ──────────────────────────
    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
    st.markdown("<div style='font-size:10px;font-weight:700;letter-spacing:1.5px;"
                "color:#4d84b8;text-transform:uppercase;margin-bottom:6px;'>Filtrar Obra</div>",
                unsafe_allow_html=True)

    obras_disponiveis = ["Todas as obras"]
    if not historico.empty:
        obras_disponiveis += sorted(historico["Obra"].dropna().unique().tolist())

    obra_selecionada = st.selectbox("", obras_disponiveis, label_visibility="collapsed")

    st.divider()
    st.markdown("""
    <div style="font-size:10px; color:#2a4d78; text-align:center; padding-top:6px;">
        VL Construtora © 2025<br>
        <span style="font-size:9px;">Atualização semanal via upload</span>
    </div>""", unsafe_allow_html=True)

# ── Aplica filtro de obra ──────────────────────────────────────────────────────
df_filtrado = historico.copy()
if obra_selecionada != "Todas as obras" and not historico.empty:
    df_filtrado = historico[historico["Obra"] == obra_selecionada].copy()


# ═════════════════════════════════════════════════════════
#  ABA: DASHBOARD
# ═════════════════════════════════════════════════════════

if menu == "📊  Dashboard":

    # ── Cabeçalho ────────────────────────────────────────
    col_h, col_btn = st.columns([6, 1])
    with col_h:
        obra_label = obra_selecionada if obra_selecionada != "Todas as obras" else "Todas as obras"
        st.markdown(f"""
        <div style="margin-bottom:6px;">
            <span style="font-size:22px;font-weight:800;color:#eaf2ff;">Dashboard</span>
            <span style="font-size:12px;color:#4d84b8;margin-left:12px;">{obra_label}</span>
        </div>
        """, unsafe_allow_html=True)

    with col_btn:
        if st.button("↻ Recalcular"):
            st.session_state.historico = calcular_colunas_derivadas(historico)
            salvar_historico(st.session_state.historico)
            st.rerun()

    ultima = datetime.now().strftime("%d/%m/%Y %H:%M")
    st.markdown(f'<div class="ts-badge">🕐 Atualizado em: <strong>{ultima}</strong></div>',
                unsafe_allow_html=True)
    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

    # ── Upload Base 1 ─────────────────────────────────────
    st.markdown('<div class="sec-title">📤 Importar Base 1</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Arraste o arquivo .xlsx exportado do ERP",
        type=["xlsx"],
        label_visibility="collapsed",
        help="O sistema identifica automaticamente as solicitações e preserva dados já preenchidos."
    )

    if uploaded:
        with st.spinner("Processando Base 1..."):
            novos_df = parse_base1(uploaded)

        if novos_df.empty:
            st.warning("⚠️ Nenhuma solicitação encontrada. Verifique se o arquivo é o relatório correto.")
        else:
            df_merged, qtd_novas = merge_com_historico(historico, novos_df)
            st.session_state.historico = df_merged
            historico = df_merged
            df_filtrado = historico if obra_selecionada == "Todas as obras" else historico[
                historico["Obra"] == obra_selecionada]
            salvar_historico(df_merged)

            if qtd_novas == 0:
                st.info("ℹ️ Todas as solicitações já constam no histórico. Nenhum registro novo adicionado.")
            else:
                st.success(f"✅ {qtd_novas} nova(s) solicitação(ões) importada(s) com sucesso!")
                st.rerun()

    st.divider()

    # ── KPIs Principais ──────────────────────────────────
    st.markdown('<div class="sec-title">📈 Indicadores</div>', unsafe_allow_html=True)

    # Conta solicitações únicas (regra: múltiplos serviços = 1 solicitação)
    total_sol   = df_filtrado["Solicitação"].nunique()
    total_obras = df_filtrado["Obra"].nunique()
    total_solic_pessoas = df_filtrado["Solicitante"].nunique()
    total_servicos = df_filtrado["Qtd Serviços"].apply(
        lambda x: int(x) if str(x).isdigit() else 0
    ).sum()

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(kpi_card(
            "Solicitações Únicas", total_sol,
            f"{total_servicos} serviços no total", "📋", "#006fd6"
        ), unsafe_allow_html=True)
    with k2:
        st.markdown(kpi_card(
            "Obras Monitoradas", total_obras,
            "obras ativas no período", "🏗️", "#0ea5e9"
        ), unsafe_allow_html=True)
    with k3:
        st.markdown(kpi_card(
            "Solicitantes", total_solic_pessoas,
            "usuários com solicitações", "👤", "#38bdf8"
        ), unsafe_allow_html=True)
    with k4:
        pendentes_prazo = int((df_filtrado["Prazo Resposta (dias)"].apply(
            lambda x: int(x) if str(x).lstrip("-").isdigit() else 0) > 30).sum())
        st.markdown(kpi_card(
            "Acima de 30 dias", pendentes_prazo,
            "solicitações com prazo estourado", "⏱️", "#f59e0b"
        ), unsafe_allow_html=True)

    st.markdown("<div style='margin-top:18px;'></div>", unsafe_allow_html=True)

    # ── Status Cards ─────────────────────────────────────
    st.markdown('<div class="sec-title">🚦 Status das Solicitações</div>', unsafe_allow_html=True)

    contagem_status = df_filtrado["Status"].value_counts() if not df_filtrado.empty else pd.Series()

    s1, s2, s3, s4 = st.columns(4)
    with s1:
        st.markdown(status_chip("Aguardando cotação",
                                int(contagem_status.get("Aguardando cotação", 0)),
                                "#f59e0b"), unsafe_allow_html=True)
    with s2:
        st.markdown(status_chip("Em andamento",
                                int(contagem_status.get("Em andamento", 0)),
                                "#3b82f6"), unsafe_allow_html=True)
    with s3:
        st.markdown(status_chip("Concluído",
                                int(contagem_status.get("Concluído", 0)),
                                "#10b981"), unsafe_allow_html=True)
    with s4:
        st.markdown(status_chip("Em atraso",
                                int(contagem_status.get("Em atraso", 0)),
                                "#ef4444"), unsafe_allow_html=True)

    st.divider()

    # ── Gráficos ─────────────────────────────────────────
    if not df_filtrado.empty:
        st.markdown('<div class="sec-title">📊 Análises Visuais</div>', unsafe_allow_html=True)

        g1, g2 = st.columns(2)

        # Gráfico 1: Solicitações por obra (barras horizontais)
        with g1:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            obras_df = (
                df_filtrado.groupby("Obra")["Solicitação"]
                .nunique()
                .reset_index()
                .rename(columns={"Solicitação": "Qtd"})
                .sort_values("Qtd", ascending=True)
            )
            fig1 = go.Figure(go.Bar(
                x=obras_df["Qtd"], y=obras_df["Obra"],
                orientation="h",
                marker=dict(color=obras_df["Qtd"],
                            colorscale=[[0, "#003d7a"], [1, "#00b0e8"]],
                            showscale=False),
                text=obras_df["Qtd"], textposition="outside",
                textfont=dict(color="#dce6f5", size=12),
            ))
            fig1.update_layout(
                title=dict(text="Solicitações por Obra", font=dict(size=13, color="#6ea8d8")),
                xaxis=dict(visible=False, showgrid=False),
                yaxis=dict(showgrid=False),
                **PLOTLY_BASE,
            )
            st.plotly_chart(fig1, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Gráfico 2: Distribuição por Solicitante (donut)
        with g2:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            solic_df = (
                df_filtrado.groupby("Solicitante")["Solicitação"]
                .nunique()
                .reset_index()
                .rename(columns={"Solicitação": "Qtd"})
            )
            fig2 = go.Figure(go.Pie(
                labels=solic_df["Solicitante"], values=solic_df["Qtd"],
                hole=0.62,
                marker=dict(colors=["#006fd6", "#00b0e8", "#50d4f8",
                                    "#0042a0", "#38bdf8", "#7dd3fc"]),
                textfont=dict(color="#dce6f5", size=11),
                hovertemplate="%{label}<br>%{value} solicitações<extra></extra>",
            ))
            fig2.update_layout(
                title=dict(text="Solicitações por Solicitante",
                           font=dict(size=13, color="#6ea8d8")),
                legend=dict(font=dict(color="#b8d0ee", size=10)),
                **PLOTLY_BASE,
            )
            st.plotly_chart(fig2, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        g3, g4 = st.columns(2)

        # Gráfico 3: Status das solicitações (barras)
        with g3:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            if not contagem_status.empty:
                cores = [COR_STATUS.get(s, "#006fd6") for s in contagem_status.index]
                fig3 = go.Figure(go.Bar(
                    x=contagem_status.index.tolist(),
                    y=contagem_status.values.tolist(),
                    marker_color=cores,
                    text=contagem_status.values.tolist(),
                    textposition="outside",
                    textfont=dict(color="#dce6f5", size=12),
                ))
                fig3.update_layout(
                    title=dict(text="Distribuição por Status",
                               font=dict(size=13, color="#6ea8d8")),
                    xaxis=dict(showgrid=False),
                    yaxis=dict(visible=False, showgrid=False),
                    **PLOTLY_BASE,
                )
                st.plotly_chart(fig3, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Gráfico 4: Evolução temporal (linha)
        with g4:
            st.markdown('<div class="chart-box">', unsafe_allow_html=True)
            df_temp = df_filtrado.copy()
            df_temp["Data_dt"] = pd.to_datetime(
                df_temp["Data solicitada"], dayfirst=True, errors="coerce")
            df_temp = df_temp.dropna(subset=["Data_dt"])

            if not df_temp.empty:
                evolucao = (
                    df_temp.groupby(df_temp["Data_dt"].dt.to_period("M"))["Solicitação"]
                    .nunique()
                    .reset_index()
                )
                evolucao["Mês"] = evolucao["Data_dt"].astype(str)
                fig4 = go.Figure(go.Scatter(
                    x=evolucao["Mês"], y=evolucao["Solicitação"],
                    mode="lines+markers",
                    line=dict(color="#006fd6", width=2.5),
                    marker=dict(color="#00b0e8", size=8),
                    fill="tozeroy",
                    fillcolor="rgba(0,111,214,0.12)",
                    hovertemplate="Mês: %{x}<br>Solicitações: %{y}<extra></extra>",
                ))
                fig4.update_layout(
                    title=dict(text="Evolução Mensal de Solicitações",
                               font=dict(size=13, color="#6ea8d8")),
                    xaxis=dict(showgrid=False),
                    yaxis=dict(showgrid=False),
                    **PLOTLY_BASE,
                )
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info("Sem dados temporais suficientes para gráfico de evolução.")
            st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.info("📂 Nenhum dado ainda. Faça o upload da Base 1 para começar.")


# ═════════════════════════════════════════════════════════
#  ABA: CONTROLE DETALHADO
# ═════════════════════════════════════════════════════════

elif menu == "📋  Controle Detalhado":

    st.markdown(f"""
    <div style="margin-bottom:16px;">
        <span style="font-size:22px;font-weight:800;color:#eaf2ff;">Controle Detalhado</span>
        <span style="font-size:12px;color:#4d84b8;margin-left:12px;">
            {obra_selecionada} · {len(df_filtrado)} solicitação(ões)
        </span>
    </div>
    """, unsafe_allow_html=True)

    if df_filtrado.empty:
        st.info("📂 Nenhum dado. Importe a Base 1 pelo Dashboard.")
    else:
        # ── Filtros extras ─────────────────────────────────
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            status_opts = ["Todos"] + sorted(df_filtrado["Status"].dropna().unique().tolist())
            f_status = st.selectbox("Status", status_opts)
        with cf2:
            solic_opts = ["Todos"] + sorted(df_filtrado["Solicitante"].dropna().unique().tolist())
            f_solic = st.selectbox("Solicitante", solic_opts)
        with cf3:
            busca = st.text_input("🔍 Buscar solicitação / serviço", placeholder="ex: 2837, cerâmica...")

        df_view = df_filtrado.copy()
        if f_status != "Todos":
            df_view = df_view[df_view["Status"] == f_status]
        if f_solic != "Todos":
            df_view = df_view[df_view["Solicitante"] == f_solic]
        if busca:
            mask = (
                df_view["Solicitação"].str.contains(busca, case=False, na=False) |
                df_view["Serviços"].str.contains(busca, case=False, na=False)
            )
            df_view = df_view[mask]

        st.markdown(
            f"<div style='font-size:11px;color:#4d84b8;margin:6px 0 10px;'>"
            f"{len(df_view)} registro(s) exibido(s)</div>",
            unsafe_allow_html=True
        )

        # ── Tabs: Visualizar | Editar ──────────────────────
        tab_view, tab_edit = st.tabs(["👁️ Visualizar", "✏️ Editar Campos Manuais"])

        with tab_view:
            colunas_exibir = [
                "Solicitação", "Data solicitada", "Obra", "Solicitante",
                "Serviços", "Qtd Serviços", "Prazo Resposta (dias)",
                "Início de Cotação", "Fim de Cotação",
                "Elaboração Contrato", "Prazo de Assinatura",
                "Contrato Assinado", "Status Cotação", "Status",
            ]
            st.dataframe(
                df_view[colunas_exibir],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Prazo Resposta (dias)": st.column_config.NumberColumn(
                        "Prazo (dias)", format="%d dias"),
                    "Qtd Serviços": st.column_config.NumberColumn("Qtd Serv."),
                    "Status": st.column_config.TextColumn("Status"),
                }
            )

        with tab_edit:
            st.markdown(
                "<div style='font-size:11px;color:#6ea8d8;margin-bottom:10px;'>"
                "Edite os campos manuais abaixo. Clique em <strong>Salvar</strong> para persistir.</div>",
                unsafe_allow_html=True
            )

            colunas_editaveis = [
                "Solicitação", "Obra", "Serviços",
                "Data Início Desejado", "Início de Cotação",
                "Contrato Assinado", "Status Cotação",
                "Observação 1", "Observação 2", "Observação 3",
            ]

            df_edit = st.data_editor(
                df_view[colunas_editaveis].copy(),
                column_config={
                    "Solicitação": st.column_config.TextColumn("Nº", disabled=True),
                    "Obra":        st.column_config.TextColumn("Obra", disabled=True),
                    "Serviços":    st.column_config.TextColumn("Serviços", disabled=True),
                    "Status Cotação": st.column_config.SelectboxColumn(
                        "Status Cotação",
                        options=[
                            "Aguardando decisão da Obra",
                            "Solicitação Incompleta",
                            "Em negociação",
                            "Em cotação",
                            "Fase de Proposta",
                            "Negociação",
                            "Fornecedor Definido",
                        ],
                    ),
                    "Contrato Assinado": st.column_config.SelectboxColumn(
                        "Contrato Assinado",
                        options=["", "Assinado", "Pendente", "Cancelado"],
                    ),
                },
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
            )

            if st.button("💾 Salvar alterações"):
                # Atualiza somente as linhas visíveis no histórico global
                idx_map = df_view.index
                for col in COLUNAS_MANUAIS:
                    if col in df_edit.columns:
                        historico.loc[idx_map, col] = df_edit[col].values

                # Recalcula derivados após edição manual
                st.session_state.historico = calcular_colunas_derivadas(historico)
                salvar_historico(st.session_state.historico)
                st.success("✅ Alterações salvas e prazos recalculados!")
                st.rerun()


# ═════════════════════════════════════════════════════════
#  ABA: EXPORTAR
# ═════════════════════════════════════════════════════════

elif menu == "📥  Exportar":

    st.markdown("""
    <div style="margin-bottom:16px;">
        <span style="font-size:22px;font-weight:800;color:#eaf2ff;">Exportar Dados</span>
        <span style="font-size:12px;color:#4d84b8;margin-left:12px;">Baixe o relatório consolidado</span>
    </div>
    """, unsafe_allow_html=True)

    # KPI de registros
    st.markdown(kpi_card(
        "Registros no histórico", len(historico),
        f"{historico['Obra'].nunique()} obra(s) · {historico['Solicitação'].nunique()} solicitação(ões)",
        "📂", "#006fd6"
    ), unsafe_allow_html=True)
    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

    col_ex1, col_ex2 = st.columns(2)

    with col_ex1:
        st.markdown('<div class="sec-title">📄 Exportar CSV</div>', unsafe_allow_html=True)
        csv_data = historico.to_csv(index=False).encode("utf-8-sig")
        nome_csv = f"vl_controle_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        st.download_button(
            "⬇️ Baixar CSV",
            data=csv_data,
            file_name=nome_csv,
            mime="text/csv",
        )

    with col_ex2:
        st.markdown('<div class="sec-title">📊 Exportar Excel (.xlsx)</div>', unsafe_allow_html=True)
        try:
            xls_path = "/tmp/vl_exportado.xlsx"
            with pd.ExcelWriter(xls_path, engine="openpyxl") as writer:
                historico.to_excel(writer, sheet_name="Controle", index=False)
                # Aba com legenda de status cotação
                legenda = pd.DataFrame({
                    "Status Cotação": [
                        "Aguardando decisão da Obra",
                        "Solicitação Incompleta",
                        "Em negociação",
                        "Em cotação",
                        "Fase de Proposta",
                        "Negociação",
                        "Fornecedor Definido",
                    ]
                })
                legenda.to_excel(writer, sheet_name="Status_Referência", index=False)

            with open(xls_path, "rb") as f:
                nome_xlsx = f"vl_controle_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    "⬇️ Baixar Excel",
                    data=f.read(),
                    file_name=nome_xlsx,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Erro ao gerar Excel: {e}. Certifique-se que o openpyxl está instalado.")

    st.divider()

    # Pré-visualização do que será exportado
    st.markdown('<div class="sec-title">👁️ Pré-visualização</div>', unsafe_allow_html=True)
    st.dataframe(historico.head(20), use_container_width=True, hide_index=True)

    st.info(
        "💡 O Excel exportado contém duas abas: **Controle** (dados completos) e "
        "**Status_Referência** (legenda dos status de cotação)."
    )
